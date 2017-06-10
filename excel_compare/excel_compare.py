# coding:utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW
from flask import Flask
from flask import abort
from flask import redirect
from flask import Blueprint, render_template, send_file
from flask import Flask, render_template, request

app = Flask(__name__)

font = Font(name=u'宋体', size=10, color=RED, bold=False)


def read_excel(file_name):
    wb2 = load_workbook(file_name)
    sheet_list = wb2.get_sheet_names()
    ws = wb2.get_sheet_by_name(sheet_list[0])
    data = {}
    for row in ws.iter_rows():
        for cell in row:
            data.update({cell.coordinate: cell.value})
    return data


def data_compare(file_model, file_new):
    model_dict = read_excel(file_model)
    new_dict = read_excel(file_new)
    difference_key = []
    for k, v in new_dict.items():
        if model_dict.get(k) != v:
            difference_key.append(k)
    return difference_key


def sign_red(file_name, keys):
    wb2 = load_workbook(file_name)
    sheet_list = wb2.get_sheet_names()
    ws = wb2.get_sheet_by_name(sheet_list[0])
    for key in keys:
        cell = ws.cell(key)
        cell.font = font
    wb2.save('tmp_file/compare_result.xlsx')


def main(file_model, file_new):
    difference_key = data_compare(file_model, file_new)
    sign_red(file_new, keys=difference_key)


@app.route('/')
def index():
    return send_file('index.html')


@app.route('/compare_excel', methods=['POST'])
def main_route():
    import base64,json
    # from cStringIO import StringIO
    data = json.loads(request.data)
    x=base64.b64decode(data['file1'][78:])
    y=base64.b64decode(data['file2'][78:])
    with open('tmp_file/model.xlsx','wb') as model_file:
        model_file.write(x)
    with open('tmp_file/to_compare.xlsx','wb') as new_file:
        new_file.write(y)
    # #Use StringIO is too slowly
    # x1,y1 = StringIO(),StringIO()
    # x1.write(x)
    # x1.seek(0)
    # y1.write(y)
    # y1.seek(0)
    main('tmp_file/model.xlsx', 'tmp_file/to_compare.xlsx')
    return json.dumps(True)


if __name__ == '__main__':
    app.run(debug=True)
